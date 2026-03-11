"""Experimental adapter for historical Arduino/MPU9255 mixed text logs."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from imuops.adapters.base import BaseAdapter
from imuops.models import SessionMetadata
from imuops.session import SessionBundle
from imuops.utils import nmea_to_decimal, read_text_lines, sampling_stats
import numpy as np
import pandas as pd
from openpyxl import load_workbook

G = 9.80665
ACC_SCALE_MPS2 = (2.0 * G) / 32768.0
GYRO_SCALE_RADS = np.deg2rad(250.0 / 32768.0)
MAG_SCALE_UT = (4912.0 / 32760.0) * 0.1


@dataclass
class LegacyContext:
    code_root: Path
    log_path: Path
    spreadsheet_path: Path | None
    extracted_gps_dir: Path | None


class LegacyArduinoAdapter(BaseAdapter):
    name = "legacy_arduino"

    @classmethod
    def detect(cls, src_path: Path) -> bool:
        if src_path.is_file() and src_path.suffix.upper() == ".TXT":
            return True
        if src_path.is_dir() and (src_path / "test_data").exists():
            return True
        if src_path.is_dir() and any(src_path.glob("*.TXT")):
            return True
        return False

    @classmethod
    def ingest(cls, src_path: Path, out_dir: Path, config: dict[str, Any]) -> SessionBundle:
        context = cls._resolve_context(src_path, config.get("session_id"))
        imu, inline_gps = cls._parse_log(context.log_path)
        sidecar_gps = cls._load_sidecar_gps(context, imu, context.log_path.stem)
        gps = inline_gps if len(inline_gps) >= len(sidecar_gps) else sidecar_gps
        sheet_row = cls._lookup_sheet_row(context.spreadsheet_path, context.log_path.stem)
        notes = []
        subject_id = None
        if sheet_row:
            notes.append(sheet_row["description"])
            maybe_subject = sheet_row["description"].split()
            if maybe_subject and maybe_subject[0].isalpha():
                subject_id = maybe_subject[0]
        stats = sampling_stats(imu["t_ms"])
        metadata = SessionMetadata(
            dataset="legacy_arduino",
            session_id=context.log_path.stem,
            source_path=str(context.log_path),
            task="pdr",
            reference_type="gps" if not gps.empty else None,
            subject_id=subject_id,
            nominal_hz=stats["nominal_hz"],
            labels_available=False,
            ground_truth_available=False,
            body_location="unknown",
            device_pose="arbitrary",
            notes=notes,
            sensors={
                "imu": True,
                "mag": bool(imu[["mx", "my", "mz"]].notna().any().any()),
                "pressure": bool(imu["pressure_pa"].notna().any()),
                "temperature": bool(imu["temp_c"].notna().any()),
                "gps": bool(not gps.empty),
            },
            extra={
                "full_scale": {
                    "acc_mps2": 2.0 * G,
                    "gyro_rads": float(np.deg2rad(250.0)),
                    "mag_uT": 491.2,
                },
                "legacy_sheet_row": sheet_row,
            },
        )
        return SessionBundle(metadata=metadata, imu=imu, gps=gps)

    @classmethod
    def _resolve_context(cls, src_path: Path, session_id: str | None) -> LegacyContext:
        src_path = src_path.expanduser().resolve()
        if src_path.is_file():
            log_path = src_path
            code_root = log_path.parent.parent if log_path.parent.name == "test_data" else log_path.parent
        elif (src_path / "test_data").exists():
            code_root = src_path
            target_id = session_id or sorted((src_path / "test_data").glob("*.TXT"))[0].stem
            log_path = code_root / "test_data" / f"{target_id}.TXT"
        else:
            code_root = src_path.parent
            target_id = session_id or sorted(src_path.glob("*.TXT"))[0].stem
            log_path = src_path / f"{target_id}.TXT"
        spreadsheet_path = code_root / "test_data" / "test_log.xlsx"
        extracted_gps_dir = code_root / "Extracted GPS data"
        return LegacyContext(
            code_root=code_root,
            log_path=log_path,
            spreadsheet_path=spreadsheet_path if spreadsheet_path.exists() else None,
            extracted_gps_dir=extracted_gps_dir if extracted_gps_dir.exists() else None,
        )

    @classmethod
    def _parse_log(cls, log_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
        imu_rows: list[dict[str, float]] = []
        gps_rows: list[dict[str, Any]] = []
        last_t_ms = 0
        for line in read_text_lines(log_path):
            stripped = line.strip()
            if not stripped or stripped.startswith("Raw mag calibration values"):
                continue
            if "GPS Data:" in stripped or stripped.startswith("$GP") or stripped.startswith("$GN"):
                gps_rows.extend(cls._parse_nmea_lines(stripped, last_t_ms))
                continue
            parts = stripped.split()
            if len(parts) < 10:
                continue
            try:
                values = [float(part) for part in parts]
            except ValueError:
                continue
            last_t_ms = int(values[0])
            imu_rows.append(
                {
                    "t_ms": last_t_ms,
                    "ax": values[1] * ACC_SCALE_MPS2,
                    "ay": values[2] * ACC_SCALE_MPS2,
                    "az": values[3] * ACC_SCALE_MPS2,
                    "gx": values[4] * GYRO_SCALE_RADS,
                    "gy": values[5] * GYRO_SCALE_RADS,
                    "gz": values[6] * GYRO_SCALE_RADS,
                    "mx": values[7] * MAG_SCALE_UT,
                    "my": values[8] * MAG_SCALE_UT,
                    "mz": values[9] * MAG_SCALE_UT,
                    "temp_c": values[10] if len(values) >= 11 else np.nan,
                    "pressure_pa": values[11] if len(values) >= 12 else np.nan,
                }
            )
        imu = pd.DataFrame.from_records(imu_rows)
        gps = pd.DataFrame.from_records(gps_rows)
        return imu, gps

    @classmethod
    def _parse_nmea_lines(cls, line: str, t_ms: int | None) -> list[dict[str, Any]]:
        rows = []
        for chunk in [part.strip().strip('"') for part in line.split("$") if part.strip()]:
            sentence = "$" + chunk
            if not sentence.startswith(("$GPRMC", "$GNRMC")):
                continue
            parts = sentence.split(",")
            if len(parts) < 7:
                continue
            valid = parts[2] == "A"
            lat = nmea_to_decimal(parts[3], parts[4]) if valid else np.nan
            lon = nmea_to_decimal(parts[5], parts[6]) if valid else np.nan
            rows.append(
                {
                    "t_ms": int(t_ms or len(rows) * 1000),
                    "lat": lat,
                    "lon": lon,
                    "valid": valid,
                    "raw_sentence": sentence,
                }
            )
        return rows

    @classmethod
    def _load_sidecar_gps(cls, context: LegacyContext, imu: pd.DataFrame, stem: str) -> pd.DataFrame:
        if context.extracted_gps_dir is None:
            return pd.DataFrame()
        candidates = [
            context.extracted_gps_dir / f"{stem}_GPS.txt",
            context.extracted_gps_dir / f"{stem.capitalize()}_GPS.txt",
            context.extracted_gps_dir / f"{stem.upper()}_GPS.txt",
        ]
        gps_lines = None
        for candidate in candidates:
            if candidate.exists():
                gps_lines = read_text_lines(candidate)
                break
        if gps_lines is None:
            return pd.DataFrame()
        rows = []
        for line in gps_lines:
            rows.extend(cls._parse_nmea_lines(line, None))
        if not rows:
            return pd.DataFrame()
        gps = pd.DataFrame.from_records(rows)
        if imu.empty:
            gps["t_ms"] = np.arange(len(gps)) * 1000
        else:
            gps["t_ms"] = np.linspace(int(imu["t_ms"].min()), int(imu["t_ms"].max()), num=len(gps)).astype(int)
        return gps

    @classmethod
    def _lookup_sheet_row(cls, sheet_path: Path | None, stem: str) -> dict[str, Any] | None:
        if sheet_path is None or not sheet_path.exists():
            return None
        wb = load_workbook(sheet_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4 and str(row[2]).upper() == stem.upper():
                return {
                    "date": row[0].isoformat() if row[0] else None,
                    "test_id": row[2],
                    "description": row[3] or "",
                    "accel_g": row[4],
                    "gyro_dps": row[5],
                    "mag_bits": row[6],
                }
        return None
